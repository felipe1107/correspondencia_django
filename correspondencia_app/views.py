from django.shortcuts           import render, redirect, get_object_or_404
from django.contrib.auth        import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic       import ListView, CreateView, UpdateView, DeleteView
from django.urls                import reverse_lazy
from django.contrib             import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator      import Paginator
from django.http                import HttpResponse
from django.db.models           import Count, Q
import csv, openpyxl
from openpyxl.utils             import get_column_letter

from .models import CorrespondenciaEntrada, DocumentManager
from .forms  import EntradaForm, DocumentManagerForm

# --- Mixins ---
class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para esa acción.")
        return redirect('correspondencia_app:dashboard_stats')

# --- Authentication ---
def login_view(request):
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user:
            login(request, user)
            messages.success(request, f"Bienvenido, {user.username}")
            return redirect('correspondencia_app:dashboard_stats')
        messages.error(request, "Usuario o contraseña incorrectos")
    return render(request, 'correspondencia_app/login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión")
    return redirect('correspondencia_app:login')

# --- Dashboard ---
@login_required
def dashboard_stats(request):
    by_gestor = (
        CorrespondenciaEntrada.objects
        .values('gestor__nombre')
        .annotate(total=Count('pk'))
        .order_by('-total')
    )
    by_estado = (
        CorrespondenciaEntrada.objects
        .values('estado')
        .annotate(total=Count('pk'))
        .order_by('-total')
    )
    return render(request, 'correspondencia_app/dashboard_stats.html', {
        'by_gestor': list(by_gestor),
        'by_estado': list(by_estado),
    })

# --- Entradas CBV ---
class EntradaListView(LoginRequiredMixin, ListView):
    model = CorrespondenciaEntrada
    template_name = 'correspondencia_app/entrada_list.html'
    context_object_name = 'entradas'
    paginate_by = 10
    ordering = ['-fecha_recepcion']

    def get_queryset(self):
        qs = super().get_queryset()
        q  = self.request.GET.get('q','')
        if q:
            qs = qs.filter(Q(asunto__icontains=q) | Q(remitente__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q','')
        return ctx

class EntradaCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = CorrespondenciaEntrada
    form_class = EntradaForm
    template_name = 'correspondencia_app/entrada_form.html'
    success_url  = reverse_lazy('correspondencia_app:entrada_list')

    def form_valid(self, form):
        messages.success(self.request, "Entrada creada correctamente.")
        return super().form_valid(form)

class EntradaUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = CorrespondenciaEntrada
    form_class = EntradaForm
    template_name = 'correspondencia_app/entrada_form.html'
    success_url  = reverse_lazy('correspondencia_app:entrada_list')

    def form_valid(self, form):
        messages.success(self.request, "Entrada actualizada correctamente.")
        return super().form_valid(form)

class EntradaDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = CorrespondenciaEntrada
    template_name = 'correspondencia_app/entrada_confirm_delete.html'
    success_url = reverse_lazy('correspondencia_app:entrada_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Entrada eliminada correctamente.")
        return super().delete(request, *args, **kwargs)

# --- Exportaciones ---
@login_required
def export_entradas_csv(request):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="entradas.csv"'
    w = csv.writer(resp)
    w.writerow(['ID','Número','Asunto','Fecha','Remitente','Destinatario','Estado','Gestor'])
    for e in CorrespondenciaEntrada.objects.order_by('-fecha_recepcion'):
        w.writerow([e.pk, e.numero_documento, e.asunto, e.fecha_recepcion,
                    e.remitente, e.destinatario, e.estado,
                    e.gestor.nombre if e.gestor else ''])
    return resp

@login_required
def export_entradas_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Entradas'
    hdr = ['ID','Número','Asunto','Fecha','Remitente','Destinatario','Estado','Gestor']
    ws.append(hdr)
    for e in CorrespondenciaEntrada.objects.order_by('-fecha_recepcion'):
        ws.append([e.pk, e.numero_documento, e.asunto, e.fecha_recepcion.strftime('%Y-%m-%d'),
                   e.remitente, e.destinatario, e.estado,
                   e.gestor.nombre if e.gestor else ''])
    for i in range(1, len(hdr)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename=entradas.xlsx'
    wb.save(resp)
    return resp

# --- Gestores ---
@login_required
@user_passes_test(lambda u: u.is_staff)
def document_manager_create(request):
    form = DocumentManagerForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Gestor creado correctamente")
            return redirect('correspondencia_app:document_manager_list')
        messages.error(request, "Corrige los errores")
    return render(request, 'correspondencia_app/document_manager_form.html', {'form': form, 'edit': False})

@login_required
@user_passes_test(lambda u: u.is_staff)
def document_manager_list(request):
    q = request.GET.get('q','')
    qs = DocumentManager.objects.order_by('nombre')
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(email__icontains=q))
    page = Paginator(qs, 10).get_page(request.GET.get('page'))
    return render(request, 'correspondencia_app/document_manager_list.html', {'gestores': page, 'q': q})

@login_required
@user_passes_test(lambda u: u.is_staff)
def document_manager_edit(request, pk):
    g = get_object_or_404(DocumentManager, pk=pk)
    form = DocumentManagerForm(request.POST or None, request.FILES or None, instance=g)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Gestor actualizado correctamente")
            return redirect('correspondencia_app:document_manager_list')
        messages.error(request, "Corrige los errores")
    return render(request, 'correspondencia_app/document_manager_form.html', {'form': form, 'edit': True})

@login_required
@user_passes_test(lambda u: u.is_staff)
def document_manager_delete(request, pk):
    g = get_object_or_404(DocumentManager, pk=pk)
    if request.method == 'POST':
        g.delete()
        messages.success(request, "Gestor eliminado correctamente")
        return redirect('correspondencia_app:document_manager_list')
    return render(request, 'correspondencia_app/document_manager_confirm_delete.html', {'gestor': g})

# --- Debug ---
@login_required
def debug_entradas(request):
    entradas = CorrespondenciaEntrada.objects.order_by('-pk')[:10]
    return render(request, 'correspondencia_app/debug_entradas.html', {'entradas': entradas})